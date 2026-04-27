from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
FONT_SIZE_ANNOTATION = 15
FONT_SIZE_AXIS_LABEL = 15
FONT_SIZE_TICK = 15
DEFAULT_TOPK_FILES = {
	1: BASE_DIR / "perron_ranking_1.csv",
	2: BASE_DIR / "perron_ranking_2.csv",
	5: BASE_DIR / "perron_ranking_5.csv",
	10: BASE_DIR / "perron_ranking_10.csv",
	20: BASE_DIR / "perron_ranking_20.csv",
	30: BASE_DIR / "perron_ranking_30.csv",
	40: BASE_DIR / "perron_ranking_40.csv",
	50: BASE_DIR / "perron_ranking_50.csv",
}


@dataclass(frozen=True)
class RankingResult:
	topk: int
	file_path: Path
	rank_by_model: dict[str, float]
	ordered_models: list[str]


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Compute pairwise Spearman correlation between model rankings and plot model rank trends across the same-level Perron tables."
	)
	for topk, default_path in DEFAULT_TOPK_FILES.items():
		parser.add_argument(
			f"--top{topk}",
			type=Path,
			default=default_path,
			help=f"Excel file containing model averages for top{topk}",
		)
	parser.add_argument(
		"--output-csv",
		type=Path,
		default=BASE_DIR / "spearman_correlation_matrix.csv",
		help="Output CSV file for the correlation matrix.",
	)
	parser.add_argument(
		"--output-png",
		type=Path,
		default=BASE_DIR / "spearman_topk_line.png",
		help="Optional line chart image output for Spearman correlation vs top50 baseline. If matplotlib is unavailable, the file is not created.",
	)
	parser.add_argument(
		"--output-pdf",
		type=Path,
		default=BASE_DIR / "Spearman_topk_line.pdf",
		help="Optional high-quality PDF output for the line chart. If matplotlib is unavailable, the file is not created.",
	)
	parser.add_argument(
		"--series-csv",
		type=Path,
		default=BASE_DIR / "spearman_topk_series.csv",
		help="Output CSV file showing the correlation series against the top10 baseline.",
	)
	return parser.parse_args()


def read_model_average_file(file_path: Path) -> list[tuple[str, float]]:
	if not file_path.is_file():
		raise FileNotFoundError(f"Input file not found: {file_path}")

	if file_path.suffix.lower() == ".csv":
		with file_path.open("r", newline="", encoding="utf-8-sig") as handle:
			reader = csv.DictReader(handle)
			if reader.fieldnames is None:
				raise ValueError(f"Empty CSV file: {file_path}")
			headers = {name.strip().lower() for name in reader.fieldnames if name is not None}
			required = {"model", "score"}
			missing = required - headers
			if missing:
				raise ValueError(f"Missing columns {sorted(missing)} in {file_path}")

			model_values: list[tuple[str, float]] = []
			for row in reader:
				model = str(row.get("model", "")).strip()
				if not model:
					continue
				value_text = str(row.get("score", "")).strip()
				if not value_text:
					continue
				model_values.append((model, float(value_text)))
	else:
		try:
			from openpyxl import load_workbook
		except Exception as exc:
			raise ImportError("openpyxl is required to read Excel input files") from exc

		workbook = load_workbook(file_path, read_only=True, data_only=True)
		if "model_average" in workbook.sheetnames:
			sheet = workbook["model_average"]
		else:
			sheet = workbook[workbook.sheetnames[0]]

		rows = list(sheet.iter_rows(values_only=True))
		if not rows:
			raise ValueError(f"Empty workbook: {file_path}")

		headers = ["" if value is None else str(value).strip() for value in rows[0]]
		index_by_header = {header: index for index, header in enumerate(headers)}
		required = {"model", "average_category_mae"}
		missing = required - set(headers)
		if missing:
			raise ValueError(f"Missing columns {sorted(missing)} in {file_path}")

		model_values = []
		for row in rows[1:]:
			if not row or all(value is None for value in row):
				continue
			model = str(row[index_by_header["model"]]).strip()
			if not model:
				continue
			value = row[index_by_header["average_category_mae"]]
			if value is None:
				continue
			model_values.append((model, float(value)))

	if not model_values:
		raise ValueError(f"No model ranking rows found in {file_path}")
	return model_values


def average_ranks(values: list[tuple[str, float]]) -> dict[str, float]:
	ordered = sorted(values, key=lambda item: (item[1], item[0]))
	ranks: dict[str, float] = {}
	position = 1
	index = 0
	while index < len(ordered):
		end = index + 1
		while end < len(ordered) and ordered[end][1] == ordered[index][1]:
			end += 1
		average_rank = (position + position + (end - index) - 1) / 2.0
		for model, _ in ordered[index:end]:
			ranks[model] = average_rank
		position += end - index
		index = end
	return ranks


def build_ranking_result(topk: int, file_path: Path) -> RankingResult:
	model_values = read_model_average_file(file_path)
	ordered = sorted(model_values, key=lambda item: (item[1], item[0]))
	rank_by_model = average_ranks(model_values)
	return RankingResult(topk=topk, file_path=file_path, rank_by_model=rank_by_model, ordered_models=[model for model, _ in ordered])


def pearson_correlation(x_values: list[float], y_values: list[float]) -> float:
	if len(x_values) != len(y_values):
		raise ValueError("Input vectors must have the same length")
	count = len(x_values)
	if count < 2:
		return math.nan

	mean_x = sum(x_values) / count
	mean_y = sum(y_values) / count
	numerator = 0.0
	denominator_x = 0.0
	denominator_y = 0.0
	for x_value, y_value in zip(x_values, y_values):
		diff_x = x_value - mean_x
		diff_y = y_value - mean_y
		numerator += diff_x * diff_y
		denominator_x += diff_x * diff_x
		denominator_y += diff_y * diff_y

	denominator = math.sqrt(denominator_x * denominator_y)
	if denominator == 0:
		return math.nan
	return numerator / denominator


def spearman_correlation(left: RankingResult, right: RankingResult) -> tuple[float, int]:
	common_models = sorted(set(left.rank_by_model) & set(right.rank_by_model))
	if len(common_models) < 2:
		return math.nan, len(common_models)
	left_ranks = [left.rank_by_model[model] for model in common_models]
	right_ranks = [right.rank_by_model[model] for model in common_models]
	return pearson_correlation(left_ranks, right_ranks), len(common_models)


def write_matrix_csv(output_path: Path, labels: list[str], matrix: list[list[float]]) -> None:
	output_path.parent.mkdir(parents=True, exist_ok=True)
	with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
		writer = csv.writer(handle)
		writer.writerow(["topk"] + labels)
		for label, row in zip(labels, matrix):
			writer.writerow([label] + ["" if math.isnan(value) else f"{value:.6f}" for value in row])


def write_rank_trends_csv(output_path: Path, ranking_results: list[RankingResult]) -> None:
	all_models = sorted(set().union(*(result.rank_by_model.keys() for result in ranking_results)))
	labels = [f"top{result.topk}" for result in ranking_results]
	output_path.parent.mkdir(parents=True, exist_ok=True)
	with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
		writer = csv.writer(handle)
		writer.writerow(["model"] + labels)
		for model in all_models:
			row = [model]
			for result in ranking_results:
				value = result.rank_by_model.get(model)
				if value is None:
					row.append("")
				else:
					row.append(f"{value:.6f}")
			writer.writerow(row)


def compute_series(ranking_results: list[RankingResult], baseline_topk: int = 50) -> list[tuple[str, float]]:
	baseline = next((result for result in ranking_results if result.topk == baseline_topk), None)
	if baseline is None:
		raise ValueError(f"Baseline top{baseline_topk} not found")
	series: list[tuple[str, float]] = []
	for result in ranking_results:
		if result.topk == baseline_topk:
			series.append((f"top{result.topk}", 1.0))
		else:
			value, _ = spearman_correlation(baseline, result)
			series.append((f"top{result.topk}", value))
	return series


def write_series_csv(output_path: Path, series: list[tuple[str, float]]) -> None:
	output_path.parent.mkdir(parents=True, exist_ok=True)
	with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
		writer = csv.writer(handle)
		writer.writerow(["topk", "spearman_correlation_vs_top50"])
		for label, value in series:
			writer.writerow([label, "" if math.isnan(value) else f"{value:.6f}"])


def try_draw_line_chart(output_png: Path, output_pdf: Path, series: list[tuple[str, float]]) -> bool:
	try:
		import matplotlib.pyplot as plt
	except Exception:
		return False

	plt.rcParams.update({
		"font.family": "Times New Roman",
		"pdf.fonttype": 42,
		"ps.fonttype": 42,
	})
	fig, ax = plt.subplots(figsize=(8.0, 5.0))
	x_labels = [label.removeprefix("top") for label, _ in series]
	x_positions = list(range(len(x_labels)))
	y_values = [value for _, value in series]
	ax.plot(x_positions, y_values, marker="o", linewidth=2.5, color="#1f77b4")
	for x_position, (label, value) in zip(x_positions, series):
		if not math.isnan(value):
			if label == "top10":
				offset_x = 0
				offset_y = 8
				vertical_alignment = "bottom"
				horizontal_alignment = "center"
			else:
				offset_y = -14 if label in {"top1", "top2", "top5"} else 8
				vertical_alignment = "top" if label in {"top1", "top2", "top5"} else "bottom"
				offset_x = 0
				horizontal_alignment = "center"
			ax.annotate(
				f"{value:.3f}",
				(x_position, value),
				textcoords="offset points",
				xytext=(offset_x, offset_y),
				ha=horizontal_alignment,
				va=vertical_alignment,
				fontsize=FONT_SIZE_ANNOTATION,
				fontname="Times New Roman",
			)
	ax.set_xticks(x_positions)
	ax.set_xticklabels(x_labels, fontname="Times New Roman")
	ax.set_ylim(0.9, 1.01)
	ax.set_xlabel("Top-k value", fontname="Times New Roman", fontsize=FONT_SIZE_AXIS_LABEL)
	ax.set_ylabel("Spearman correlation", fontname="Times New Roman", fontsize=FONT_SIZE_AXIS_LABEL)
	#ax.set_title("Spearman correlation vs top50 baseline", fontname="Times New Roman")
	ax.axhline(0.99, color="#d62728", linewidth=1.5)
	ax.tick_params(axis="both", labelsize=FONT_SIZE_TICK)
	ax.spines["top"].set_visible(False)
	ax.spines["right"].set_visible(False)
	ax.spines["left"].set_linewidth(1.5)
	ax.spines["bottom"].set_linewidth(1.5)
	ax.annotate(
		"",
		xy=(1.02, 0.0),
		xytext=(0.0, 0.0),
		xycoords="axes fraction",
		textcoords="axes fraction",
		arrowprops={"arrowstyle": "-|>", "lw": 1.5, "color": "black"},
		annotation_clip=False,
	)
	ax.annotate(
		"",
		xy=(0.0, 1.02),
		xytext=(0.0, 0.0),
		xycoords="axes fraction",
		textcoords="axes fraction",
		arrowprops={"arrowstyle": "-|>", "lw": 1.5, "color": "black"},
		annotation_clip=False,
	)
	for tick in ax.get_yticklabels():
		tick.set_fontname("Times New Roman")
	for tick in ax.get_xticklabels():
		tick.set_fontname("Times New Roman")
	fig.tight_layout()
	output_png.parent.mkdir(parents=True, exist_ok=True)
	fig.savefig(output_png, dpi=300, bbox_inches="tight")
	fig.savefig(output_pdf, format="pdf", bbox_inches="tight")
	plt.close(fig)
	return True


def main() -> int:
	args = parse_args()

	ranking_results: list[RankingResult] = []
	for topk in sorted(DEFAULT_TOPK_FILES):
		file_path = getattr(args, f"top{topk}")
		ranking_results.append(build_ranking_result(topk, file_path))

	labels = [f"top{result.topk}" for result in ranking_results]
	matrix: list[list[float]] = []
	common_counts: dict[tuple[str, str], int] = {}

	for left in ranking_results:
		row: list[float] = []
		for right in ranking_results:
			value, common_count = spearman_correlation(left, right)
			row.append(value)
			common_counts[(f"top{left.topk}", f"top{right.topk}")] = common_count
		matrix.append(row)

	series = compute_series(ranking_results, baseline_topk=50)
	write_matrix_csv(args.output_csv, labels, matrix)
	write_series_csv(args.series_csv, series)
	chart_written = try_draw_line_chart(args.output_png, args.output_pdf, series)

	print("Spearman correlation matrix:")
	print("\t" + "\t".join(labels))
	for label, row in zip(labels, matrix):
		formatted = ["nan" if math.isnan(value) else f"{value:.6f}" for value in row]
		print(label + "\t" + "\t".join(formatted))

	print(f"CSV saved to {args.output_csv}")
	print(f"Series saved to {args.series_csv}")
	if chart_written:
		print(f"Line chart saved to {args.output_png}")
		print(f"PDF chart saved to {args.output_pdf}")
	else:
		print("matplotlib not available; line chart was not created")

	for result in ranking_results:
		print(f"{result.file_path.name}: {len(result.rank_by_model)} models ranked")

	return 0


if __name__ == "__main__":
	raise SystemExit(main())
