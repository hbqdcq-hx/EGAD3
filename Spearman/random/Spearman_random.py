from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
FONT_SIZE_ANNOTATION = 15
FONT_SIZE_AXIS_LABEL = 15
FONT_SIZE_TICK = 15
BASELINE_SEED = 8409
DEFAULT_SEED_FILES = {
	4738: BASE_DIR / "mae_seed_4738.xlsx",
	5183: BASE_DIR / "mae_seed_5183.xlsx",
	2947: BASE_DIR / "mae_seed_2947.xlsx",
	1362: BASE_DIR / "mae_seed_1362.xlsx",
	8409: BASE_DIR / "mae_seed_8409.xlsx",
	3651: BASE_DIR / "mae_seed_3651.xlsx",
	4024: BASE_DIR / "mae_seed_4024.xlsx",
	915: BASE_DIR / "mae_seed_915.xlsx",
}


@dataclass(frozen=True)
class RankingResult:
	seed: int
	file_path: Path
	rank_by_model: dict[str, float]
	ordered_models: list[str]


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Compute pairwise Spearman correlation between model rankings from mae_seed Excel summaries and plot the correlation trend."
	)
	for seed, default_path in DEFAULT_SEED_FILES.items():
		parser.add_argument(
			f"--seed-{seed}",
			dest=f"seed_{seed}",
			type=Path,
			default=default_path,
			help=f"Excel file containing model averages for seed {seed}",
		)
	parser.add_argument(
		"--output-csv",
		type=Path,
		default=BASE_DIR / "spearman_seed_correlation_matrix.csv",
		help="Output CSV file for the correlation matrix.",
	)
	parser.add_argument(
		"--output-png",
		type=Path,
		default=BASE_DIR / "spearman_seed_line.png",
		help="Optional line chart image output for Spearman correlation vs the baseline seed. If matplotlib is unavailable, the file is not created.",
	)
	parser.add_argument(
		"--output-pdf",
		type=Path,
		default=BASE_DIR / "Spearman_seed_line.pdf",
		help="Optional high-quality PDF output for the line chart. If matplotlib is unavailable, the file is not created.",
	)
	parser.add_argument(
		"--series-csv",
		type=Path,
		default=BASE_DIR / "spearman_seed_series.csv",
		help="Output CSV file showing the correlation series against the baseline seed.",
	)
	return parser.parse_args()


def _normalize_headers(fieldnames: list[str]) -> dict[str, str]:
	return {name.strip().lower(): name for name in fieldnames if name is not None}


def _pick_column(headers: dict[str, str], candidates: list[str]) -> str | None:
	for candidate in candidates:
		key = candidate.strip().lower()
		if key in headers:
			return headers[key]
	return None


def read_model_average_file(file_path: Path) -> list[tuple[str, float]]:
	if not file_path.is_file():
		raise FileNotFoundError(f"Input file not found: {file_path}")

	if file_path.suffix.lower() == ".csv":
		with file_path.open("r", newline="", encoding="utf-8-sig") as handle:
			reader = csv.DictReader(handle)
			if reader.fieldnames is None:
				raise ValueError(f"Empty CSV file: {file_path}")
			headers = _normalize_headers([name for name in reader.fieldnames if name is not None])
			model_column = _pick_column(headers, ["model", "模型"])
			value_column = _pick_column(headers, ["average_category_mae", "average_mae", "平均mae", "mae", "score", "分数"])
			if model_column is None or value_column is None:
				raise ValueError(f"Missing columns in {file_path}: expected model/模型 and average_category_mae/平均mae-like value column")

			model_values: list[tuple[str, float]] = []
			for row in reader:
				model = str(row.get(model_column, "")).strip()
				if not model:
					continue
				value_text = str(row.get(value_column, "")).strip()
				if not value_text:
					continue
				model_values.append((model, float(value_text)))
	else:
		try:
			from openpyxl import load_workbook
		except Exception as exc:
			raise ImportError("openpyxl is required to read Excel input files") from exc

		workbook = load_workbook(file_path, read_only=True, data_only=True)
		if "summary" in workbook.sheetnames:
			sheet = workbook["summary"]
		elif "model_average" in workbook.sheetnames:
			sheet = workbook["model_average"]
		else:
			sheet = workbook[workbook.sheetnames[0]]

		rows = list(sheet.iter_rows(values_only=True))
		if not rows:
			raise ValueError(f"Empty workbook: {file_path}")

		headers = ["" if value is None else str(value).strip() for value in rows[0]]
		headers_map = _normalize_headers(headers)
		model_column = _pick_column(headers_map, ["model", "模型"])
		value_column = _pick_column(headers_map, ["average_category_mae", "average_mae", "平均mae", "mae", "score", "分数"])
		if model_column is None or value_column is None:
			raise ValueError(f"Missing columns in {file_path}: expected model/模型 and average_category_mae/平均mae-like value column")

		model_index = headers.index(model_column)
		value_index = headers.index(value_column)
		model_values = []
		for row in rows[1:]:
			if not row or all(value is None for value in row):
				continue
			model = "" if row[model_index] is None else str(row[model_index]).strip()
			if not model:
				continue
			value = row[value_index]
			if value is None:
				continue
			model_values.append((model, float(value)))

	if not model_values:
		raise ValueError(f"No model ranking rows found in {file_path}")
	return model_values


def average_ranks(values: list[tuple[str, float]]) -> dict[str, float]:
	ordered = sorted(values, key=lambda item: (item[1], item[0]))
	ranks: dict[str, float] = {}
	position = 1
	index = 0
	while index < len(ordered):
		end = index + 1
		while end < len(ordered) and ordered[end][1] == ordered[index][1]:
			end += 1
		average_rank = (position + position + (end - index) - 1) / 2.0
		for model, _ in ordered[index:end]:
			ranks[model] = average_rank
		position += end - index
		index = end
	return ranks


def build_ranking_result(seed: int, file_path: Path) -> RankingResult:
	model_values = read_model_average_file(file_path)
	ordered = sorted(model_values, key=lambda item: (item[1], item[0]))
	rank_by_model = average_ranks(model_values)
	return RankingResult(seed=seed, file_path=file_path, rank_by_model=rank_by_model, ordered_models=[model for model, _ in ordered])


def pearson_correlation(x_values: list[float], y_values: list[float]) -> float:
	if len(x_values) != len(y_values):
		raise ValueError("Input vectors must have the same length")
	count = len(x_values)
	if count < 2:
		return math.nan

	mean_x = sum(x_values) / count
	mean_y = sum(y_values) / count
	numerator = 0.0
	denominator_x = 0.0
	denominator_y = 0.0
	for x_value, y_value in zip(x_values, y_values):
		diff_x = x_value - mean_x
		diff_y = y_value - mean_y
		numerator += diff_x * diff_y
		denominator_x += diff_x * diff_x
		denominator_y += diff_y * diff_y

	denominator = math.sqrt(denominator_x * denominator_y)
	if denominator == 0:
		return math.nan
	return numerator / denominator


def spearman_correlation(left: RankingResult, right: RankingResult) -> tuple[float, int]:
	common_models = sorted(set(left.rank_by_model) & set(right.rank_by_model))
	if len(common_models) < 2:
		return math.nan, len(common_models)
	left_ranks = [left.rank_by_model[model] for model in common_models]
	right_ranks = [right.rank_by_model[model] for model in common_models]
	return pearson_correlation(left_ranks, right_ranks), len(common_models)


def write_matrix_csv(output_path: Path, labels: list[str], matrix: list[list[float]]) -> None:
	output_path.parent.mkdir(parents=True, exist_ok=True)
	with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
		writer = csv.writer(handle)
		writer.writerow(["seed"] + labels)
		for label, row in zip(labels, matrix):
			writer.writerow([label] + ["" if math.isnan(value) else f"{value:.6f}" for value in row])


def write_rank_trends_csv(output_path: Path, ranking_results: list[RankingResult]) -> None:
	all_models = sorted(set().union(*(result.rank_by_model.keys() for result in ranking_results)))
	labels = [f"seed{result.seed}" for result in ranking_results]
	output_path.parent.mkdir(parents=True, exist_ok=True)
	with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
		writer = csv.writer(handle)
		writer.writerow(["model"] + labels)
		for model in all_models:
			row = [model]
			for result in ranking_results:
				value = result.rank_by_model.get(model)
				if value is None:
					row.append("")
				else:
					row.append(f"{value:.6f}")
			writer.writerow(row)


def compute_series(ranking_results: list[RankingResult], baseline_seed: int = BASELINE_SEED) -> list[tuple[str, float]]:
	baseline = next((result for result in ranking_results if result.seed == baseline_seed), None)
	if baseline is None:
		raise ValueError(f"Baseline seed {baseline_seed} not found")
	series: list[tuple[str, float]] = []
	for result in ranking_results:
		if result.seed == baseline_seed:
			series.append((f"seed{result.seed}", 1.0))
		else:
			value, _ = spearman_correlation(baseline, result)
			series.append((f"seed{result.seed}", value))
	return series


def write_series_csv(output_path: Path, series: list[tuple[str, float]]) -> None:
	output_path.parent.mkdir(parents=True, exist_ok=True)
	with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
		writer = csv.writer(handle)
		writer.writerow(["seed", "spearman_correlation_vs_baseline_seed"])
		for label, value in series:
			writer.writerow([label, "" if math.isnan(value) else f"{value:.6f}"])


def try_draw_line_chart(output_png: Path, output_pdf: Path, series: list[tuple[str, float]]) -> bool:
	try:
		import matplotlib.pyplot as plt
	except Exception:
		return False

	plt.rcParams.update({
		"font.family": "Times New Roman",
		"pdf.fonttype": 42,
		"ps.fonttype": 42,
	})
	fig, ax = plt.subplots(figsize=(8.0, 5.0))
	x_labels = [label.removeprefix("seed") for label, _ in series]
	x_positions = list(range(len(x_labels)))
	y_values = [value for _, value in series]
	ax.plot(x_positions, y_values, marker="o", linewidth=2.5, color="#1f77b4")
	ax.set_ylim(0.0, 1.06)
	ax.set_yticks([0.0, 0.2, 0.4, 0.6, 0.8, 1.0])
	for x_position, (label, value) in zip(x_positions, series):
		if not math.isnan(value):
			if label == f"seed{BASELINE_SEED}":
				offset_x = 0
				offset_y = 8
				vertical_alignment = "bottom"
				horizontal_alignment = "center"
			else:
				offset_y = -14 if label in {"seed20220421", "seed20230421", "seed20240421"} else 8
				vertical_alignment = "top" if label in {"seed20220421", "seed20230421", "seed20240421"} else "bottom"
				offset_x = 0
				horizontal_alignment = "center"
			ax.annotate(
				f"{value:.3f}",
				(x_position, value),
				textcoords="offset points",
				xytext=(offset_x, offset_y),
				ha=horizontal_alignment,
				va=vertical_alignment,
				fontsize=FONT_SIZE_ANNOTATION,
				fontname="Times New Roman",
			)
	ax.set_xticks(x_positions)
	ax.set_xticklabels(x_labels, fontname="Times New Roman")
	ax.set_xlabel("Seed", fontname="Times New Roman", fontsize=FONT_SIZE_AXIS_LABEL)
	ax.set_ylabel("Spearman correlation", fontname="Times New Roman", fontsize=FONT_SIZE_AXIS_LABEL)
	#ax.set_title("Spearman correlation vs baseline seed", fontname="Times New Roman")
	ax.axhline(1.0, color="#d62728", linewidth=1.5)
	ax.tick_params(axis="both", labelsize=FONT_SIZE_TICK)
	ax.spines["top"].set_visible(False)
	ax.spines["right"].set_visible(False)
	ax.spines["left"].set_linewidth(1.5)
	ax.spines["bottom"].set_linewidth(1.5)
	ax.annotate(
		"",
		xy=(1.02, 0.0),
		xytext=(0.0, 0.0),
		xycoords="axes fraction",
		textcoords="axes fraction",
		arrowprops={"arrowstyle": "-|>", "lw": 1.5, "color": "black"},
		annotation_clip=False,
	)
	ax.annotate(
		"",
		xy=(0.0, 1.02),
		xytext=(0.0, 0.0),
		xycoords="axes fraction",
		textcoords="axes fraction",
		arrowprops={"arrowstyle": "-|>", "lw": 1.5, "color": "black"},
		annotation_clip=False,
	)
	for tick in ax.get_yticklabels():
		tick.set_fontname("Times New Roman")
	for tick in ax.get_xticklabels():
		tick.set_fontname("Times New Roman")
	fig.tight_layout()
	output_png.parent.mkdir(parents=True, exist_ok=True)
	fig.savefig(output_png, dpi=300, bbox_inches="tight")
	fig.savefig(output_pdf, format="pdf", bbox_inches="tight")
	plt.close(fig)
	return True


def main() -> int:
	args = parse_args()

	ranking_results: list[RankingResult] = []
	for seed in sorted(DEFAULT_SEED_FILES):
		file_path = getattr(args, f"seed_{seed}")
		ranking_results.append(build_ranking_result(seed, file_path))

	labels = [f"seed{result.seed}" for result in ranking_results]
	matrix: list[list[float]] = []
	common_counts: dict[tuple[str, str], int] = {}

	for left in ranking_results:
		row: list[float] = []
		for right in ranking_results:
			value, common_count = spearman_correlation(left, right)
			row.append(value)
			common_counts[(f"seed{left.seed}", f"seed{right.seed}")] = common_count
		matrix.append(row)

	series = compute_series(ranking_results, baseline_seed=BASELINE_SEED)
	write_matrix_csv(args.output_csv, labels, matrix)
	write_rank_trends_csv(BASE_DIR / "spearman_seed_rank_trends.csv", ranking_results)
	write_series_csv(args.series_csv, series)
	chart_written = try_draw_line_chart(args.output_png, args.output_pdf, series)

	print("Spearman correlation matrix:")
	print("	" + "	".join(labels))
	for label, row in zip(labels, matrix):
		formatted = ["nan" if math.isnan(value) else f"{value:.6f}" for value in row]
		print(label + "	" + "	".join(formatted))

	print(f"CSV saved to {args.output_csv}")
	print(f"Series saved to {args.series_csv}")
	if chart_written:
		print(f"Line chart saved to {args.output_png}")
		print(f"PDF chart saved to {args.output_pdf}")
	else:
		print("matplotlib not available; line chart was not created")

	for result in ranking_results:
		print(f"{result.file_path.name}: {len(result.rank_by_model)} models ranked")

	return 0


if __name__ == "__main__":
	raise SystemExit(main())
