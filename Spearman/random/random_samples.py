#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import random
from collections import defaultdict
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_SCORE_DIR = BASE_DIR / "data_score"
DEFAULT_LABEL_DIR = BASE_DIR / "label"
DEFAULT_OUTPUT_DIR = BASE_DIR / "random_mae_outputs"
DEFAULT_RANDOM_SEED = 4024
DEFAULT_SAMPLE_N = 10

CATEGORY_ORDER = [
    "bottle",
    "cable",
    "capsule",
    "carpet",
    "grid",
    "hazelnut",
    "leather",
    "metal_nut",
    "pill",
    "screw",
    "tile",
    "toothbrush",
    "transistor",
    "wood",
    "zipper",
]

LABEL_FILE_MAP = {
    "bottle": "bottle_label.csv",
    "cable": "cable_label.csv",
    "capsule": "capsule_label.csv",
    "carpet": "carpet_label.csv",
    "grid": "grid_label.csv",
    "hazelnut": "hazelnut_label.csv",
    "leather": "leather_label.csv",
    "metal_nut": "metal_nut_label.csv",
    "pill": "pill_label.csv",
    "screw": "screw_label.csv",
    "tile": "tile_label.csv",
    "toothbrush": "toothbrush_label.csv",
    "transistor": "transistor_label.csv",
    "wood": "wood_label.csv",
    "zipper": "zipper_label.csv",
}


@contextmanager
def open_csv(path: Path):
	last_error: Exception | None = None
	for encoding in ("utf-8-sig", "gbk", "utf-8"):
		try:
			handle = path.open("r", encoding=encoding, newline="")
			try:
				handle.read(1)
				handle.seek(0)
			except UnicodeDecodeError as exc:
				handle.close()
				last_error = exc
				continue
			else:
				yield handle
				handle.close()
				return
		except FileNotFoundError:
			raise
		except Exception as exc:
			last_error = exc
	if last_error is not None:
		raise last_error


def natural_sort_key(value: str) -> list[object]:
	parts: list[object] = []
	current = ""
	is_digit = None
	for char in value:
		char_is_digit = char.isdigit()
		if is_digit is None:
			current = char
			is_digit = char_is_digit
			continue
		if char_is_digit == is_digit:
			current += char
		else:
			parts.append(int(current) if is_digit else current.lower())
			current = char
			is_digit = char_is_digit
	if current:
		parts.append(int(current) if is_digit else current.lower())
	return parts


def sanitize_sheet_name(value: str) -> str:
	invalid_chars = set('[]:*?/\\')
	text = "".join("_" if char in invalid_chars else char for char in value.strip())
	text = text[:31].strip()
	return text or "sheet"


def normalize_name(value: str) -> str:
	text = value.strip().replace("\\", "/")
	parts = [part for part in text.split("/") if part]
	if not parts:
		return Path(text).stem
	if len(parts) == 1:
		return Path(parts[0]).stem
	return "_".join(Path(part).stem for part in parts)


def label_lookup_key(image_name: str) -> str:
	return Path(image_name).stem.split("_")[-1]


def resolve_label_file(category: str, label_dir: Path) -> Path:
	label_path = label_dir / LABEL_FILE_MAP.get(category, f"{category}_label.csv")
	if not label_path.is_file():
		raise FileNotFoundError(f"找不到 label 文件: {label_path}")
	return label_path


def resolve_score_file(model_dir: Path, category: str) -> Path:
	for file_name in (f"anomaly_scores_{category}.csv", f"anomaly_scores_mvtec_{category}.csv"):
		candidate = model_dir / file_name
		if candidate.is_file():
			return candidate
	raise FileNotFoundError(f"找不到 score 文件: model={model_dir.name}, category={category}")


def load_label_rows(category: str, label_dir: Path) -> dict[str, dict[str, object]]:
	label_path = resolve_label_file(category, label_dir)
	with open_csv(label_path) as handle:
		reader = csv.DictReader(handle)
		if reader.fieldnames is None:
			raise ValueError(f"空文件或缺少表头: {label_path}")

		normalized = {name.strip().lower(): name for name in reader.fieldnames}
		required = {"filename", "label"}
		missing = required - set(normalized)
		if missing:
			raise ValueError(f"label 文件缺少列 {sorted(missing)}: {label_path}")

		label_rows: dict[str, dict[str, object]] = {}
		for row in reader:
			filename = str(row.get(normalized["filename"], "")).strip()
			if not filename:
				continue
			label_rows[Path(filename).stem] = {
				"filename": filename,
				"label": float(str(row.get(normalized["label"], "")).strip()),
			}
		return label_rows


def load_score_map(score_path: Path) -> dict[str, float]:
	with open_csv(score_path) as handle:
		reader = csv.DictReader(handle)
		if reader.fieldnames is None:
			raise ValueError(f"空文件或缺少表头: {score_path}")

		normalized = {name.strip().lower(): name for name in reader.fieldnames}
		name_column = None
		for key in ("file_name", "file_path", "filename", "path"):
			if key in normalized:
				name_column = normalized[key]
				break
		if name_column is None:
			raise ValueError(f"score 文件缺少文件名列: {score_path}")

		score_column = normalized.get("anomaly_score")
		if score_column is None:
			raise ValueError(f"score 文件缺少 Anomaly_Score 列: {score_path}")

		scores: dict[str, float] = {}
		for row in reader:
			image_name = normalize_name(str(row.get(name_column, "")))
			if not image_name:
				continue
			scores[label_lookup_key(image_name)] = float(str(row.get(score_column, "")).strip())
		return scores


def get_model_dirs(data_score_dir: Path) -> list[Path]:
	if not data_score_dir.is_dir():
		raise FileNotFoundError(f"找不到 data_score 目录: {data_score_dir}")
	return sorted(path for path in data_score_dir.iterdir() if path.is_dir())


def build_model_score_cache(model_dirs: list[Path]) -> dict[str, dict[str, dict[str, float]]]:
	cache: dict[str, dict[str, dict[str, float]]] = {}
	for model_dir in model_dirs:
		category_scores: dict[str, dict[str, float]] = {}
		for category in CATEGORY_ORDER:
			score_path = resolve_score_file(model_dir, category)
			category_scores[category] = load_score_map(score_path)
		cache[model_dir.name] = category_scores
	return cache


def build_sample_keys(
	label_dir: Path,
	model_score_cache: dict[str, dict[str, dict[str, float]]],
	sample_n: int,
	rng: random.Random,
) -> dict[str, list[str]]:
	sampled_keys: dict[str, list[str]] = {}
	for category in CATEGORY_ORDER:
		label_rows = load_label_rows(category, label_dir)
		if not label_rows:
			raise ValueError(f"标签文件没有可用样本: {category}")

		common_keys = set(label_rows)
		for model_name, category_scores in model_score_cache.items():
			common_keys &= set(category_scores.get(category, {}))

		available_keys = sorted(common_keys, key=natural_sort_key)
		if len(available_keys) < sample_n:
			raise ValueError(
				f"类别 {category} 可用样本不足 {sample_n} 个: 只有 {len(available_keys)} 个公共样本"
			)

		selected_keys = rng.sample(available_keys, sample_n)
		sampled_keys[category] = sorted(selected_keys, key=natural_sort_key)

	return sampled_keys


def build_sample_mean_mae(
	label_dir: Path,
	model_score_cache: dict[str, dict[str, dict[str, float]]],
	sampled_keys: dict[str, list[str]],
) -> dict[tuple[str, str], float]:
	sample_mean_mae: dict[tuple[str, str], float] = {}
	model_names = list(model_score_cache)

	for category in CATEGORY_ORDER:
		label_rows = load_label_rows(category, label_dir)
		for key in sampled_keys[category]:
			label_value = float(label_rows[key]["label"])
			mae_values = [abs(model_score_cache[model_name][category][key] - label_value) for model_name in model_names]
			sample_mean_mae[(category, key)] = sum(mae_values) / len(mae_values)

	return sample_mean_mae


def build_rows_for_model(
	model_name: str,
	model_scores: dict[str, dict[str, float]],
	label_dir: Path,
	sampled_keys: dict[str, list[str]],
	sample_mean_mae: dict[tuple[str, str], float],
) -> tuple[list[dict[str, object]], float | None]:
	rows: list[dict[str, object]] = []
	mae_values: list[float] = []

	for category in CATEGORY_ORDER:
		label_rows = load_label_rows(category, label_dir)
		category_scores = model_scores[category]
		for key in sampled_keys[category]:
			label_row = label_rows[key]
			label_value = float(label_row["label"])
			score_value = float(category_scores[key])
			mae_value = abs(score_value - label_value)
			rows.append(
				{
					"类别": category,
					"名称": label_row["filename"],
					"标签": label_value,
					"分数": score_value,
					"mae": mae_value,
					"所有模型平均mae": sample_mean_mae[(category, key)],
				}
			)
			mae_values.append(mae_value)

	overall_mean_mae = sum(mae_values) / len(mae_values) if mae_values else None
	for row in rows:
		row["模型"] = model_name

	return rows, overall_mean_mae


def write_workbook(
	data_score_dir: Path,
	label_dir: Path,
	output_dir: Path,
	seed: int,
	sample_n: int,
) -> Path:
	model_dirs = get_model_dirs(data_score_dir)
	if not model_dirs:
		raise FileNotFoundError(f"data_score 目录下没有模型子目录: {data_score_dir}")

	model_score_cache = build_model_score_cache(model_dirs)
	rng = random.Random(seed)
	sampled_keys = build_sample_keys(label_dir, model_score_cache, sample_n, rng)
	sample_mean_mae = build_sample_mean_mae(label_dir, model_score_cache, sampled_keys)

	workbook = Workbook()
	workbook.remove(workbook.active)

	model_summary_rows: list[dict[str, object]] = []
	for model_dir in model_dirs:
		model_name = model_dir.name
		rows, overall_mean_mae = build_rows_for_model(
			model_name,
			model_score_cache[model_name],
			label_dir,
			sampled_keys,
			sample_mean_mae,
		)
		model_summary_rows.append(
			{
				"模型": model_name,
				"样本数": len(rows),
				"平均mae": overall_mean_mae,
			}
		)

		sheet = workbook.create_sheet(title=sanitize_sheet_name(model_name))
		headers = ["模型", "类别", "名称", "标签", "分数", "mae", "所有模型平均mae"]
		for col_index, header in enumerate(headers, start=1):
			cell = sheet.cell(row=1, column=col_index, value=header)
			cell.font = Font(bold=True)

		for row_index, row in enumerate(rows, start=2):
			for col_index, header in enumerate(headers, start=1):
				sheet.cell(row=row_index, column=col_index, value=row.get(header, ""))

		sheet.freeze_panes = "A2"

	summary_sheet = workbook.create_sheet(title="summary", index=0)
	summary_headers = ["模型", "样本数", "平均mae"]
	for col_index, header in enumerate(summary_headers, start=1):
		cell = summary_sheet.cell(row=1, column=col_index, value=header)
		cell.font = Font(bold=True)

	for row_index, row in enumerate(model_summary_rows, start=2):
		for col_index, header in enumerate(summary_headers, start=1):
			summary_sheet.cell(row=row_index, column=col_index, value=row.get(header, ""))

	summary_sheet.freeze_panes = "A2"
	output_dir.mkdir(parents=True, exist_ok=True)
	output_path = output_dir / f"mae_seed_{seed}.xlsx"
	workbook.save(output_path)
	return output_path


def build_parser() -> argparse.ArgumentParser:
	parser = argparse.ArgumentParser(description="按类别随机抽样并基于 data_score 计算 MAE，输出每个模型一个 Excel sheet。")
	parser.add_argument("--data-score-dir", type=Path, default=DEFAULT_DATA_SCORE_DIR, help="包含模型子目录的 data_score 目录")
	parser.add_argument("--label-dir", type=Path, default=DEFAULT_LABEL_DIR, help="label 目录")
	parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="输出目录")
	parser.add_argument("--seed", type=int, default=DEFAULT_RANDOM_SEED, help=f"随机种子，默认 {DEFAULT_RANDOM_SEED}")
	parser.add_argument("--sample-n", type=int, default=DEFAULT_SAMPLE_N, help=f"每个类别随机抽样数量，默认 {DEFAULT_SAMPLE_N}")
	return parser


def main() -> None:
	args = build_parser().parse_args()
	output_path = write_workbook(
		data_score_dir=args.data_score_dir,
		label_dir=args.label_dir,
		output_dir=args.output_dir,
		seed=args.seed,
		sample_n=args.sample_n,
	)
	print(f"已生成 Excel: {output_path}")


if __name__ == "__main__":
	main()
