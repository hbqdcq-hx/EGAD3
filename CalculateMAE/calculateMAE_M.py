from __future__ import annotations

import argparse
import csv
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "mvtec_score"
DEFAULT_OUTPUT_FILE = BASE_DIR / "mvtec_score_mae_summary.xlsx"


@contextmanager
def open_csv(path: Path):
	last_error: Exception | None = None
	for encoding in ("utf-8-sig", "gbk", "utf-8"):
		try:
			handle = path.open("r", encoding=encoding, newline="")
			try:
				handle.read(1)
				handle.seek(0)
			except UnicodeDecodeError as exc:
				handle.close()
				last_error = exc
				continue
			else:
				yield handle
				handle.close()
				return
		except FileNotFoundError:
			raise
		except Exception as exc:
			last_error = exc
	if last_error is not None:
		raise last_error


def normalize_name(value: str) -> str:
	return Path(str(value).strip().replace("\\", "/")).stem.lower()


def detect_columns(fieldnames: list[str]) -> tuple[str, str]:
	normalized = {name.strip().lower(): name for name in fieldnames}
	name_column = None
	for key in ("file_name", "file_path", "filename", "path"):
		if key in normalized:
			name_column = normalized[key]
			break
	if name_column is None:
		raise ValueError("CSV 缺少文件名列")

	score_column = None
	for key in ("anomaly_score", "score"):
		if key in normalized:
			score_column = normalized[key]
			break
	if score_column is None:
		for name in fieldnames:
			text = name.strip().lower()
			if text.endswith("score"):
				score_column = name
				break
	if score_column is None:
		raise ValueError("CSV 缺少分数列")

	return name_column, score_column


def collect_model_dirs(data_dir: Path) -> list[Path]:
	if not data_dir.is_dir():
		raise FileNotFoundError(f"找不到 mvtec_score 目录: {data_dir}")
	return sorted(path for path in data_dir.iterdir() if path.is_dir())


def collect_score_files(model_dir: Path) -> list[Path]:
	return sorted(path for path in model_dir.rglob("*.csv") if path.is_file())


def compute_model_mae(model_dir: Path) -> dict[str, object]:
	file_count = 0
	sample_count = 0
	good_count = 0
	defect_count = 0
	total_mae_sum = 0.0
	good_mae_sum = 0.0
	defect_mae_sum = 0.0
	rows: list[dict[str, object]] = []

	for score_file in collect_score_files(model_dir):
		file_count += 1
		with open_csv(score_file) as handle:
			reader = csv.DictReader(handle)
			if reader.fieldnames is None:
				continue
			name_column, score_column = detect_columns(reader.fieldnames)

			for row in reader:
				file_name = str(row.get(name_column, "")).strip()
				if not file_name:
					continue
				score_text = str(row.get(score_column, "")).strip()
				if not score_text:
					continue

				label = 0 if "good" in file_name.lower() else 1
				score = float(score_text)
				mae = abs(score - label)
				sample_count += 1
				total_mae_sum += mae
				if label == 0:
					good_count += 1
					good_mae_sum += mae
				else:
					defect_count += 1
					defect_mae_sum += mae
				rows.append(
					{
						"source_file": score_file.name,
						"sample_name": file_name,
						"label": label,
						"anomaly_score": score,
						"mae": mae,
					}
				)

	good_mae_mean = good_mae_sum / good_count if good_count else float("nan")
	defect_mae_mean = defect_mae_sum / defect_count if defect_count else float("nan")
	average_mae = total_mae_sum / sample_count if sample_count else float("nan")
	return {
		"model": model_dir.name,
		"file_count": file_count,
		"sample_count": sample_count,
		"good_count": good_count,
		"defect_count": defect_count,
		"total_mae_sum": total_mae_sum,
		"good_mae_mean": good_mae_mean,
		"defect_mae_mean": defect_mae_mean,
		"average_mae": average_mae,
		"rows": rows,
	}


def write_workbook(output_file: Path, model_results: list[dict[str, object]]) -> None:
	output_file.parent.mkdir(parents=True, exist_ok=True)
	workbook = Workbook()
	default_sheet = workbook.active
	default_sheet.title = "summary"

	summary_headers = [
		"model",
		"file_count",
		"sample_count",
		"good_count",
		"defect_count",
		"total_mae_sum",
		"good_mae_mean",
		"defect_mae_mean",
		"average_mae",
	]
	for col_index, column_name in enumerate(summary_headers, start=1):
		cell = default_sheet.cell(row=1, column=col_index, value=column_name)
		cell.font = Font(bold=True)

	for row_index, row in enumerate(model_results, start=2):
		for col_index, column_name in enumerate(summary_headers, start=1):
			default_sheet.cell(row=row_index, column=col_index, value=row.get(column_name, ""))

	if model_results:
		total_mae_sum = sum(float(row["total_mae_sum"]) for row in model_results)
		total_sample_count = sum(int(row["sample_count"]) for row in model_results)
		good_count = sum(int(row["good_count"]) for row in model_results)
		defect_count = sum(int(row["defect_count"]) for row in model_results)
		good_mae_sum = sum(float(row["good_mae_mean"]) * int(row["good_count"]) for row in model_results if int(row["good_count"]))
		defect_mae_sum = sum(float(row["defect_mae_mean"]) * int(row["defect_count"]) for row in model_results if int(row["defect_count"]))
		good_mae_mean = good_mae_sum / good_count if good_count else float("nan")
		defect_mae_mean = defect_mae_sum / defect_count if defect_count else float("nan")
		overall_average_mae = total_mae_sum / total_sample_count if total_sample_count else float("nan")
		overall_row = {
			"model": "overall",
			"file_count": sum(int(row["file_count"]) for row in model_results),
			"sample_count": total_sample_count,
			"good_count": good_count,
			"defect_count": defect_count,
			"total_mae_sum": total_mae_sum,
			"good_mae_mean": good_mae_mean,
			"defect_mae_mean": defect_mae_mean,
			"average_mae": overall_average_mae,
		}
		row_index = len(model_results) + 2
		for col_index, column_name in enumerate(summary_headers, start=1):
			cell = default_sheet.cell(row=row_index, column=col_index, value=overall_row.get(column_name, ""))
			if column_name == "model":
				cell.font = Font(bold=True)

	default_sheet.freeze_panes = "A2"

	for model_result in model_results:
		model_name = str(model_result["model"])
		sheet_name = model_name[:31] if model_name else "sheet"
		sheet = workbook.create_sheet(title=sheet_name)
		headers = ["source_file", "sample_name", "label", "anomaly_score", "mae"]
		for col_index, column_name in enumerate(headers, start=1):
			cell = sheet.cell(row=1, column=col_index, value=column_name)
			cell.font = Font(bold=True)

		for row_index, row in enumerate(model_result["rows"], start=2):
			for col_index, column_name in enumerate(headers, start=1):
				sheet.cell(row=row_index, column=col_index, value=row.get(column_name, ""))

		metrics_start_row = len(model_result["rows"]) + 3
		metrics = [
			("good_mae_mean", model_result["good_mae_mean"]),
			("defect_mae_mean", model_result["defect_mae_mean"]),
			("average_mae", model_result["average_mae"]),
		]
		for offset, (metric_name, metric_value) in enumerate(metrics):
			row = metrics_start_row + offset
			sheet.cell(row=row, column=1, value=metric_name).font = Font(bold=True)
			sheet.cell(row=row, column=2, value=metric_value)
		sheet.freeze_panes = "A2"

	workbook.save(output_file)


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="计算 mvtec_score 下每个模型的平均 MAE，并输出 Excel 汇总。")
	parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR, help="包含各模型 score CSV 的目录")
	parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE, help="输出 Excel 文件")
	return parser.parse_args()


def main() -> None:
	args = parse_args()
	model_dirs = collect_model_dirs(args.data_dir)
	if not model_dirs:
		raise FileNotFoundError(f"在目录中没有找到模型子目录: {args.data_dir}")

	rows = [compute_model_mae(model_dir) for model_dir in model_dirs]
	write_workbook(args.output_file, rows)

	for row in rows:
		print(
			f"{row['model']}: label_0_mae_mean={row['good_mae_mean']:.6f}, "
			f"label_1_mae_mean={row['defect_mae_mean']:.6f}, average_mae={row['average_mae']:.6f}, "
			f"samples={row['sample_count']}, good={row['good_count']}, defect={row['defect_count']}"
		)
	print(f"已输出 Excel 文件: {args.output_file}")


if __name__ == "__main__":
	main()
