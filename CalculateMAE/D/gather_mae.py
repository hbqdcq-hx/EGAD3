from __future__ import annotations

import argparse
import csv
import math
from collections import defaultdict
from io import StringIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_MERGED_CSV = BASE_DIR / "pairwise_results_1_merged_dedup.csv"
DEFAULT_LABEL_DIR = BASE_DIR / "label"
DEFAULT_DATA_SCORE_DIR = BASE_DIR / "data_score"
DEFAULT_OUTPUT = BASE_DIR / "mae_by_model_category_1.xlsx"

CATEGORY_ORDER = [
	"bottle",
	"cable",
	"capsule",
	"carpet",
	"grid",
	"hazelnut",
	"leather",
	"metal_nut",
	"pill",
	"screw",
	"tile",
	"toothbrush",
	"transistor",
	"wood",
	"zipper",
]

LABEL_FILE_MAP = {
	"bottle": "bottle_label.csv",
	"cable": "cable_label.csv",
	"capsule": "capsule_label.csv",
	"carpet": "carpet_label.csv",
	"grid": "grid_label.csv",
	"hazelnut": "hazelnet_label.csv",
	"leather": "leather_label.csv",
	"metal_nut": "metal_nut.csv",
	"pill": "pill_label.csv",
	"screw": "screw_label.csv",
	"tile": "tile_label.csv",
	"toothbrush": "toothbrush_label.csv",
	"transistor": "transistor_label.csv",
	"wood": "wood_label.csv",
	"zipper": "zipper_label.csv",
}


def safe_float(value: object) -> float:
	if value is None:
		return float("nan")
	text = str(value).strip()
	if not text:
		return float("nan")
	try:
		return float(text)
	except ValueError:
		return float("nan")


def normalize_name(value: object) -> str:
	text = str(value).strip().replace("\\", "/")
	if not text:
		return ""
	parts = [part for part in text.split("/") if part]
	if not parts:
		return Path(text).stem
	if len(parts) == 1:
		return Path(parts[0]).stem
	return "_".join(Path(part).stem for part in parts)


def label_lookup_key(image_name: str) -> str:
	return Path(image_name).stem.split("_")[-1]


def category_sort_key(category: str) -> tuple[int, str]:
	if category in CATEGORY_ORDER:
		return (CATEGORY_ORDER.index(category), category)
	return (len(CATEGORY_ORDER), category)


def sanitize_sheet_name(name: str) -> str:
	sheet_name = name.replace("/", "_").replace("\\", "_")
	if not sheet_name:
		sheet_name = "sheet"
	return sheet_name[:31]


def read_csv_with_fallback(csv_path: Path) -> csv.DictReader:
	last_error: Exception | None = None
	for encoding in ("utf-8-sig", "utf-8", "gbk", "cp936"):
		try:
			text = csv_path.read_text(encoding=encoding)
		except UnicodeDecodeError as exc:
			last_error = exc
			continue
		reader = csv.DictReader(StringIO(text))
		if reader.fieldnames is not None:
			return reader
	if last_error is not None:
		raise UnicodeDecodeError("utf-8", b"", 0, 1, f"无法识别文件编码: {csv_path}") from last_error
	raise UnicodeDecodeError("utf-8", b"", 0, 1, f"无法识别文件编码: {csv_path}")


def load_merged_rows(merged_csv: Path) -> list[dict[str, str]]:
	if not merged_csv.is_file():
		raise FileNotFoundError(f"找不到 merged CSV: {merged_csv}")

	reader = read_csv_with_fallback(merged_csv)
	if reader.fieldnames is None:
		raise ValueError(f"空文件或缺少表头: {merged_csv}")

	required = {"类别", "图片文件名"}
	missing = required - set(reader.fieldnames)
	if missing:
		raise ValueError(f"merged CSV 缺少列 {sorted(missing)}: {merged_csv}")

	rows_by_key: dict[tuple[str, str], dict[str, str]] = {}
	for row in reader:
		category = str(row.get("类别", "")).strip()
		image_name = str(row.get("图片文件名", "")).strip()
		if not category or not image_name:
			continue
		key = (category, image_name)
		rows_by_key[key] = {
			"来源文件": str(row.get("来源文件", "")).strip(),
			"类别": category,
			"图片文件名": image_name,
			"类别内排名": str(row.get("类别内排名", "")).strip(),
			"分数差异": str(row.get("分数差异", "")).strip(),
			"模型1": str(row.get("模型1", "")).strip(),
			"模型1分数": str(row.get("模型1分数", "")).strip(),
			"模型2": str(row.get("模型2", "")).strip(),
			"模型2分数": str(row.get("模型2分数", "")).strip(),
			"来源文件数": str(row.get("来源文件数", "")).strip(),
			"来源文件列表": str(row.get("来源文件列表", "")).strip(),
		}

	rows = list(rows_by_key.values())
	rows.sort(key=lambda item: (category_sort_key(item["类别"]), item["图片文件名"]))
	return rows


def resolve_label_file(category: str, label_dir: Path) -> Path:
	file_name = LABEL_FILE_MAP.get(category, f"{category}_label.csv")
	label_path = label_dir / file_name
	if not label_path.is_file():
		raise FileNotFoundError(f"找不到 label 文件: {label_path}")
	return label_path


def load_label_map(category: str, label_dir: Path) -> dict[str, int]:
	label_path = resolve_label_file(category, label_dir)
	reader = read_csv_with_fallback(label_path)
	if reader.fieldnames is None:
		raise ValueError(f"空文件或缺少表头: {label_path}")
	headers = {name.strip().lower(): name for name in reader.fieldnames}
	if "filename" not in headers or "label" not in headers:
		raise ValueError(f"label 文件缺少 filename/label 列: {label_path}")

	filename_col = headers["filename"]
	label_col = headers["label"]
	label_map: dict[str, int] = {}
	for row in reader:
		filename = str(row.get(filename_col, "")).strip()
		if not filename:
			continue
		label_key = Path(filename).stem
		label_map[label_key] = int(float(str(row.get(label_col, "")).strip()))
	return label_map


def resolve_score_file(model_dir: Path, category: str) -> Path:
	candidates = [
		model_dir / f"anomaly_scores_{category}.csv",
		model_dir / f"anomaly_scores_mvtec_{category}.csv",
	]
	for candidate in candidates:
		if candidate.is_file():
			return candidate
	raise FileNotFoundError(f"找不到 score 文件: model={model_dir.name}, category={category}")


def load_score_map(score_path: Path) -> dict[str, float]:
	reader = read_csv_with_fallback(score_path)
	if reader.fieldnames is None:
		raise ValueError(f"空文件或缺少表头: {score_path}")

	headers = {name.strip().lower(): name for name in reader.fieldnames}
	name_col = None
	for key in ("file_name", "file_path", "filename", "path"):
		if key in headers:
			name_col = headers[key]
			break
	if name_col is None:
		raise ValueError(f"score 文件缺少文件名列: {score_path}")

	score_col = None
	for key in ("anomaly_score", "score"):
		if key in headers:
			score_col = headers[key]
			break
	if score_col is None:
		raise ValueError(f"score 文件缺少 Anomaly_Score 列: {score_path}")

	scores: dict[str, float] = {}
	for row in reader:
		image_name = normalize_name(row.get(name_col, ""))
		if not image_name:
			continue
		scores[image_name] = safe_float(row.get(score_col, ""))
	return scores


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="基于 pairwise_results_20_merged_dedup.csv 生成每个模型的 MAE 汇总表。")
	parser.add_argument("--merged-csv", type=Path, default=DEFAULT_MERGED_CSV, help="pairwise_results_20_merged_dedup.csv 路径")
	parser.add_argument("--label-dir", type=Path, default=DEFAULT_LABEL_DIR, help="label 目录")
	parser.add_argument("--data-score-dir", type=Path, default=DEFAULT_DATA_SCORE_DIR, help="data_score 目录")
	parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出 xlsx 文件")
	return parser.parse_args()


def collect_rows(
	merged_csv: Path,
	label_dir: Path,
	data_score_dir: Path,
) -> tuple[list[str], list[dict[str, object]], dict[str, list[dict[str, object]]]]:
	rows = load_merged_rows(merged_csv)
	if not label_dir.is_dir():
		raise FileNotFoundError(f"label 目录不存在: {label_dir}")
	if not data_score_dir.is_dir():
		raise FileNotFoundError(f"data_score 目录不存在: {data_score_dir}")

	model_names = sorted(
		path.name for path in data_score_dir.iterdir() if path.is_dir() and not path.name.startswith("~$")
	)
	if not model_names:
		raise FileNotFoundError(f"在 {data_score_dir} 下没有找到模型目录")

	unique_categories = sorted({row["类别"] for row in rows}, key=category_sort_key)
	label_cache: dict[str, dict[str, int]] = {}
	score_cache: dict[tuple[str, str], dict[str, float]] = {}

	for model_name in model_names:
		model_dir = data_score_dir / model_name
		for category in unique_categories:
			score_cache[(model_name, category)] = load_score_map(resolve_score_file(model_dir, category))

	model_records: dict[str, list[dict[str, object]]] = defaultdict(list)
	summary_totals: dict[tuple[str, str], float] = defaultdict(float)
	summary_counts: dict[tuple[str, str], int] = defaultdict(int)

	for row in rows:
		category = row["类别"]
		image_name = row["图片文件名"]
		image_key = label_lookup_key(image_name)

		if category not in label_cache:
			label_cache[category] = load_label_map(category, label_dir)
		label_map = label_cache[category]
		if image_key not in label_map:
			raise ValueError(f"找不到对应 label: category={category}, image={image_name}")

		label_value = float(label_map[image_key])
		base_row = {
			"来源文件": row["来源文件"],
			"类别": category,
			"图片文件名": image_name,
			"类别内排名": row["类别内排名"],
			"分数差异": row["分数差异"],
			"模型1": row["模型1"],
			"模型1分数": row["模型1分数"],
			"模型2": row["模型2"],
			"模型2分数": row["模型2分数"],
			"来源文件数": row["来源文件数"],
			"来源文件列表": row["来源文件列表"],
		}

		for model_name in model_names:
			score_map = score_cache[(model_name, category)]
			score_value = score_map.get(normalize_name(image_name))
			if score_value is None or math.isnan(score_value):
				raise ValueError(f"找不到 score: model={model_name}, category={category}, image={image_name}")

			abs_error = abs(score_value - label_value)
			summary_totals[(model_name, category)] += abs_error
			summary_counts[(model_name, category)] += 1

			model_records[model_name].append(
				{
					**base_row,
					"model": model_name,
					"score": score_value,
					"label": label_value,
					"mae": abs_error,
				}
			)

	summary_rows: list[dict[str, object]] = []
	for model_name in model_names:
		for category in CATEGORY_ORDER:
			count = summary_counts.get((model_name, category), 0)
			if not count:
				continue
			mae_sum = summary_totals[(model_name, category)]
			summary_rows.append(
				{
					"model": model_name,
					"category": category,
					"sample_count": count,
					"category_mae": mae_sum / count,
				}
			)

	for model_name in model_names:
		model_records[model_name].sort(key=lambda item: (category_sort_key(str(item["类别"])), str(item["图片文件名"])))

	return model_names, summary_rows, model_records


def write_workbook(
	model_names: list[str],
	summary_rows: list[dict[str, object]],
	model_records: dict[str, list[dict[str, object]]],
	output_path: Path,
) -> None:
	workbook = Workbook()
	workbook.remove(workbook.active)

	summary_sheet = workbook.create_sheet(title="summary")
	summary_headers = ["model", "category", "sample_count", "category_mae"]
	for col_index, header in enumerate(summary_headers, start=1):
		cell = summary_sheet.cell(row=1, column=col_index, value=header)
		cell.font = Font(bold=True)

	for row_index, row in enumerate(summary_rows, start=2):
		for col_index, header in enumerate(summary_headers, start=1):
			summary_sheet.cell(row=row_index, column=col_index, value=row.get(header, ""))
	summary_sheet.freeze_panes = "A2"

	model_headers = [
		"model",
		"类别",
		"图片文件名",
		"类别内排名",
		"分数差异",
		"score",
		"label",
		"mae",
		"来源文件",
		"模型1",
		"模型1分数",
		"模型2",
		"模型2分数",
		"来源文件数",
		"来源文件列表",
	]

	for model_name in model_names:
		rows = model_records.get(model_name, [])
		sheet = workbook.create_sheet(title=sanitize_sheet_name(model_name))
		sheet["A1"] = "model"
		sheet["B1"] = model_name
		sheet["D1"] = "sample_count"
		sheet["E1"] = len(rows)
		sheet["G1"] = "average_mae"
		sheet["H1"] = sum(float(item["mae"]) for item in rows) / len(rows) if rows else None
		for cell in (sheet["A1"], sheet["D1"], sheet["G1"]):
			cell.font = Font(bold=True)
		for col_index, header in enumerate(model_headers, start=1):
			cell = sheet.cell(row=3, column=col_index, value=header)
			cell.font = Font(bold=True)
		for row_index, row in enumerate(rows, start=4):
			for col_index, header in enumerate(model_headers, start=1):
				sheet.cell(row=row_index, column=col_index, value=row.get(header, ""))
		sheet.freeze_panes = "A4"

	output_path.parent.mkdir(parents=True, exist_ok=True)
	workbook.save(output_path)


def main() -> None:
	args = parse_args()
	model_names, summary_rows, model_records = collect_rows(args.merged_csv, args.label_dir, args.data_score_dir)
	write_workbook(model_names, summary_rows, model_records, args.output)
	print(f"已汇总 {len(model_names)} 个模型，样本数 {len(next(iter(model_records.values()), [])) if model_records else 0}，输出文件: {args.output}")


if __name__ == "__main__":
	main()