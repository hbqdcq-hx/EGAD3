from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "pairwise_results_1"
DEFAULT_LABEL_DIR = BASE_DIR / "label"
DEFAULT_OUTPUT_DIR = BASE_DIR / "pairwise_results_1_mae"

LABEL_FILE_MAP = {
	"bottle": "bottle_label.csv",
	"cable": "cable_label.csv",
	"capsule": "capsule_label.csv",
	"carpet": "carpet_label.csv",
	"grid": "grid_label.csv",
	"hazelnut": "hazelnut_label.csv",
	"leather": "leather_label.csv",
	"metal_nut": "metal_nut_label.csv",
	"pill": "pill_label.csv",
	"screw": "screw_label.csv",
	"tile": "tile_label.csv",
	"toothbrush": "toothbrush_label.csv",
	"transistor": "transistor_label.csv",
	"wood": "wood_label.csv",
	"zipper": "zipper_label.csv",
}

CATEGORY_ORDER = [
	"bottle",
	"cable",
	"capsule",
	"carpet",
	"grid",
	"hazelnut",
	"leather",
	"metal_nut",
	"pill",
	"screw",
	"tile",
	"toothbrush",
	"transistor",
	"wood",
	"zipper",
]


@contextmanager
def open_csv(path: Path):
	last_error: Exception | None = None
	for encoding in ("utf-8-sig", "gbk", "utf-8"):
		try:
			handle = path.open("r", encoding=encoding, newline="")
			try:
				handle.read(1)
				handle.seek(0)
			except UnicodeDecodeError as exc:
				handle.close()
				last_error = exc
				continue
			else:
				yield handle
				handle.close()
				return
		except FileNotFoundError:
			raise
		except Exception as exc:
			last_error = exc
	if last_error is not None:
		raise last_error


def safe_float(value: str | None) -> float:
	if value is None:
		return float("nan")
	text = value.strip()
	if not text:
		return float("nan")
	return float(text)


def normalize_name(value: str) -> str:
	text = value.strip().replace("\\", "/")
	parts = [part for part in text.split("/") if part]
	if not parts:
		return Path(text).stem
	if len(parts) == 1:
		return Path(parts[0]).stem
	return "_".join(Path(part).stem for part in parts)


def label_lookup_key(image_name: str) -> str:
	return Path(image_name).stem.split("_")[-1]


def model_name_from_score_column(score_column: str) -> str:
	text = score_column.strip()
	lower_text = text.lower()
	for suffix in ("分数", "score"):
		if lower_text.endswith(suffix):
			return text[: -len(suffix)].strip(" _-.")
	return text


def collect_pairwise_files(data_dir: Path) -> list[Path]:
	if data_dir.is_file():
		return [data_dir] if data_dir.suffix.lower() == ".csv" else []
	if not data_dir.is_dir():
		raise FileNotFoundError(f"pairwise 目录不存在: {data_dir}")
	return sorted(path for path in data_dir.glob("*.csv") if path.is_file())


def resolve_score_file(model_dir: Path, category: str) -> Path:
	candidates = [
		model_dir / f"anomaly_scores_{category}.csv",
		model_dir / f"anomaly_scores_mvtec_{category}.csv",
	]
	for candidate in candidates:
		if candidate.is_file():
			return candidate
	raise FileNotFoundError(f"找不到 score 文件: model={model_dir.name}, category={category}")


def resolve_label_file(category: str, label_dir: Path) -> Path:
	file_name = LABEL_FILE_MAP.get(category, f"{category}_label.csv")
	label_path = label_dir / file_name
	if not label_path.is_file():
		raise FileNotFoundError(f"找不到 label 文件: {label_path}")
	return label_path


def load_label_map(category: str, label_dir: Path) -> dict[str, int]:
	label_path = resolve_label_file(category, label_dir)
	with open_csv(label_path) as handle:
		reader = csv.DictReader(handle)
		if reader.fieldnames is None:
			raise ValueError(f"空文件或缺少表头: {label_path}")
		normalized = {name.strip().lower(): name for name in reader.fieldnames}
		required = {"filename", "label"}
		missing = required - set(normalized)
		if missing:
			raise ValueError(f"label 文件缺少列 {sorted(missing)}: {label_path}")

		label_map: dict[str, int] = {}
		for row in reader:
			filename = str(row.get(normalized["filename"], "")).strip()
			if not filename:
				continue
			label_key = Path(filename).stem
			label_map[label_key] = int(float(str(row.get(normalized["label"], "")).strip()))
	return label_map


def load_score_map(score_path: Path) -> dict[str, float]:
	with open_csv(score_path) as handle:
		reader = csv.DictReader(handle)
		if reader.fieldnames is None:
			raise ValueError(f"空文件或缺少表头: {score_path}")

		normalized = {name.strip().lower(): name for name in reader.fieldnames}
		name_column = None
		for key in ("file_name", "file_path", "filename", "path"):
			if key in normalized:
				name_column = normalized[key]
				break
		if name_column is None:
			raise ValueError(f"score 文件缺少文件名列: {score_path}")

		score_column = normalized.get("anomaly_score")
		if score_column is None:
			raise ValueError(f"score 文件缺少 Anomaly_Score 列: {score_path}")

		scores: dict[str, float] = {}
		for row in reader:
			image_name = normalize_name(str(row.get(name_column, "")))
			if not image_name:
				continue
			scores[image_name] = float(str(row.get(score_column, "")).strip())
		return scores


def detect_pairwise_columns(fieldnames: list[str]) -> tuple[str, str, list[str]]:
	normalized = {name.strip().lower(): name for name in fieldnames}
	category_column = None
	for key in ("类别", "category"):
		if key in normalized:
			category_column = normalized[key]
			break
	if category_column is None:
		raise ValueError("pairwise 表缺少类别列")

	image_column = None
	for key in ("图片文件名", "image_name", "filename", "file_name"):
		if key in normalized:
			image_column = normalized[key]
			break
	if image_column is None:
		raise ValueError("pairwise 表缺少图片文件名列")

	score_columns: list[str] = []
	for name in fieldnames:
		text = name.strip().lower()
		if name in {category_column, image_column}:
			continue
		if text in {"类别内排名", "分数差异", "rank", "difference", "diff"}:
			continue
		if text.endswith("分数") or text.endswith("score"):
			score_columns.append(name)

	if not score_columns:
		raise ValueError("pairwise 表中没有找到模型分数列")

	return category_column, image_column, score_columns


def compute_pairwise_rows(pairwise_path: Path, label_dir: Path) -> tuple[list[str], list[str], dict[str, list[dict[str, object]]]]:
	with open_csv(pairwise_path) as handle:
		reader = csv.DictReader(handle)
		if reader.fieldnames is None:
			raise ValueError(f"空文件或缺少表头: {pairwise_path}")

		category_column, image_column, score_columns = detect_pairwise_columns(reader.fieldnames)
		original_headers = list(reader.fieldnames)
		model_names = [model_name_from_score_column(column) for column in score_columns]
		label_cache: dict[str, dict[str, int]] = {}
		missing_labels: list[str] = []
		sheet_rows: dict[str, list[dict[str, object]]] = defaultdict(list)

		for row in reader:
			category = str(row.get(category_column, "")).strip()
			image_name = str(row.get(image_column, "")).strip()
			if not category or not image_name:
				continue

			if category not in label_cache:
				label_cache[category] = load_label_map(category, label_dir)
			label_map = label_cache[category]
			image_key = label_lookup_key(image_name)
			if image_key not in label_map:
				missing_labels.append(f"{category}/{image_name}")
				continue

			label_value = float(label_map[image_key])
			base_row = {header: row.get(header, "") for header in original_headers}
			for score_column in score_columns:
				score_text = str(row.get(score_column, "")).strip()
				if not score_text:
					raise ValueError(f"分数列为空: file={pairwise_path.name}, column={score_column}, image={image_name}")
				score_value = float(score_text)
				model_name = model_name_from_score_column(score_column)
				row_with_metrics = dict(base_row)
				row_with_metrics["label"] = label_value
				row_with_metrics["mae"] = abs(score_value - label_value)
				sheet_rows[model_name].append(row_with_metrics)

		if missing_labels:
			raise ValueError(
				f"找不到对应 label: file={pairwise_path.name}, examples={missing_labels[:5]}"
			)

		return original_headers, model_names, sheet_rows


def build_summary_rows(model_names: list[str], sheet_rows: dict[str, list[dict[str, object]]]) -> list[dict[str, object]]:
	row: dict[str, object] = {"type": "overall", "name": "overall"}
	for model_name in model_names:
		model_rows = sheet_rows.get(model_name, [])
		total_mae = sum(float(item["mae"]) for item in model_rows)
		row[f"{model_name}_mae"] = total_mae / len(model_rows) if model_rows else None
	row["sample_count"] = len(next(iter(sheet_rows.values()))) if sheet_rows else 0
	return [row]


def write_pairwise_workbook(pairwise_path: Path, label_dir: Path, output_dir: Path) -> Path:
	original_headers, model_names, sheet_rows = compute_pairwise_rows(pairwise_path, label_dir)
	workbook = Workbook()
	default_sheet = workbook.active
	workbook.remove(default_sheet)

	output_path = output_dir / f"{pairwise_path.stem}.xlsx"
	summary_sheet = workbook.create_sheet(title="summary")
	summary_headers = ["type", "name", "sample_count", *[f"{model_name}_mae" for model_name in model_names]]
	for col_index, column_name in enumerate(summary_headers, start=1):
		cell = summary_sheet.cell(row=1, column=col_index, value=column_name)
		cell.font = Font(bold=True)

	summary_rows = build_summary_rows(model_names, sheet_rows)
	for row_index, row in enumerate(summary_rows, start=2):
		for col_index, column_name in enumerate(summary_headers, start=1):
			summary_sheet.cell(row=row_index, column=col_index, value=row.get(column_name, ""))
	
	summary_sheet.freeze_panes = "A2"

	for model_name in sorted(sheet_rows):
		sheet_name = model_name[:31] if model_name else "sheet"
		sheet = workbook.create_sheet(title=sheet_name)
		headers = original_headers + ["label", "mae"]
		for col_index, column_name in enumerate(headers, start=1):
			cell = sheet.cell(row=1, column=col_index, value=column_name)
			cell.font = Font(bold=True)

		for row_index, row in enumerate(sheet_rows[model_name], start=2):
			for col_index, column_name in enumerate(headers, start=1):
				sheet.cell(row=row_index, column=col_index, value=row.get(column_name, ""))

		sheet.freeze_panes = "A2"

	output_dir.mkdir(parents=True, exist_ok=True)
	workbook.save(output_path)
	return output_path


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="计算 pairwise_results_1 中每个模型分数与 label 的 MAE。")
	parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR, help="包含 pairwise CSV 的目录")
	parser.add_argument("--label-dir", type=Path, default=DEFAULT_LABEL_DIR, help="包含 label CSV 的目录")
	parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="输出 Excel 目录")
	return parser.parse_args()


def main() -> None:
	args = parse_args()
	pairwise_files = collect_pairwise_files(args.data_dir)
	if not pairwise_files:
		raise FileNotFoundError(f"在目录中没有找到 pairwise CSV: {args.data_dir}")
	output_files: list[Path] = []
	for pairwise_file in pairwise_files:
		output_files.append(write_pairwise_workbook(pairwise_file, args.label_dir, args.output_dir))

	print(f"已生成 {len(output_files)} 个 Excel 文件，输出目录: {args.output_dir}")


if __name__ == "__main__":
	main()
